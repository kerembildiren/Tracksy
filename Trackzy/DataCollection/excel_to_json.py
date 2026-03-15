"""
Import artists from the Excel file (artists_raw.xlsx) back into artists_raw.json.
Use this after you have edited the Excel file (add/delete artists, change cells).

Usage:
  python excel_to_json.py                          # output/artists_raw.xlsx -> output/artists_raw.json
  python excel_to_json.py path/to/edited.xlsx       # custom Excel -> output/artists_raw.json
  python excel_to_json.py path/to/edited.xlsx path/to/artists_raw.json  # custom paths

Expects the same columns as produced by json_to_excel.py; sheet name "Artists".
Genre(s) column: use semicolon, comma, or slash to separate multiple (e.g. "Rap; Pop" or "Arabesk/Rap"). Empty = no genre.
"""
import json
import os
import re

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl

# Must match json_to_excel.py (order = original JSON key order)
COLUMNS = [
    ("id", "ID"),
    ("name", "Name"),
    ("gender", "Gender"),
    ("genres", "Genre(s)"),
    ("popularity", "Popularity"),
    ("debut", "Debut"),
    ("nationality", "Nationality"),
    ("group_size", "Group size"),
    ("spotify_monthly_streams", "Spotify monthly streams"),
    ("image_url", "Image URL"),
    ("top_track_name", "Top track name"),
    ("top_track_uri", "Top track URI"),
    ("top_track_id", "Top track ID"),
]

GENRE_SEP = "; "


def parse_cell(value, key):
    """Convert Excel cell value to JSON-friendly value."""
    empty = value is None or (isinstance(value, str) and value.strip() == "")
    if empty:
        if key in ("popularity", "debut", "spotify_monthly_streams"):
            return None
        if key == "group_size":
            return None
        if key == "genres":
            return []
        return None if key in ("image_url", "top_track_name", "top_track_uri", "top_track_id") else ""
    if key == "genres":
        s = str(value).strip()
        if not s:
            return []
        parts = [p.strip() for p in re.split(r"[;,/]", s) if p.strip()]
        return parts
    if key in ("popularity", "spotify_monthly_streams"):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None
    if key == "group_size":
        v = str(value).strip().lower()
        if v in ("1", "person", "solo"):
            return 1
        if v in ("group", "band", "orchestra", "choir"):
            return "group"
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return "group" if v else None
    if key == "debut":
        s = str(value).strip()
        if not s:
            return None
        m = re.search(r"\d{4}", s)
        return m.group(0) if m else s
    return str(value).strip()


def import_excel_to_json(excel_path, json_path=None):
    if json_path is None:
        json_path = os.path.join(os.path.dirname(excel_path), "artists_raw.json")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "Artists" not in wb.sheetnames:
        raise SystemExit(f"Sheet 'Artists' not found in {excel_path}")

    ws = wb["Artists"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise SystemExit("Excel sheet is empty.")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    key_by_col = {}
    for col_idx, excel_header in enumerate(header):
        for key, col_header in COLUMNS:
            if col_header == excel_header:
                key_by_col[col_idx] = key
                break

    artists = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        artist = {}
        for col_idx, key in key_by_col.items():
            cell_val = row[col_idx] if col_idx < len(row) else None
            val = parse_cell(cell_val, key)
            # Use None for empty optional fields to match original JSON
            if val == "" and key in ("nationality", "gender"):
                artist[key] = None
            else:
                artist[key] = val
        if artist.get("name") or artist.get("id"):
            # Ensure key order matches original JSON
            ordered = {}
            for key, _ in COLUMNS:
                ordered[key] = artist.get(key)
            artists.append(ordered)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(artists, f, indent=2, ensure_ascii=False)

    print(f"Imported {len(artists)} artists from {excel_path} -> {json_path}")
    return json_path


if __name__ == "__main__":
    import sys
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_excel = os.path.join(script_dir, "output", "artists_raw.xlsx")
    default_json = os.path.join(script_dir, "output", "artists_raw.json")
    excel_path = sys.argv[1] if len(sys.argv) > 1 else default_excel
    json_path = sys.argv[2] if len(sys.argv) > 2 else default_json
    import_excel_to_json(excel_path, json_path)
