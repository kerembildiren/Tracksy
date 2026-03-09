"""
1. Remove empty rows from artists_raw.xlsx (no ID and no Name).
2. Renumber Popularity as 1, 2, 3, ... in row order.
3. Save Excel and update artists_raw.json.
"""
import os
import sys

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

# Match json_to_excel column order
HEADERS = ["ID", "Name", "Gender", "Genre(s)", "Followers", "Popularity", "Debut",
           "Nationality", "Last.fm listeners", "Last.fm play count", "Group size"]

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "output", "artists_raw.xlsx")
JSON_PATH = os.path.join(SCRIPT_DIR, "output", "artists_raw.json")


def is_empty_row(row, id_col, name_col):
    id_val = row[id_col] if id_col < len(row) else None
    name_val = row[name_col] if name_col < len(row) else None
    id_ok = id_val is not None and str(id_val).strip() != ""
    name_ok = name_val is not None and str(name_val).strip() != ""
    return not id_ok and not name_ok


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Artists"]

    # Find column indices from header (row 1)
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_index = {str(h).strip(): i for i, h in enumerate(header_row) if h}
    id_col = col_index.get("ID", 0)
    name_col = col_index.get("Name", 1)
    pop_col = col_index.get("Popularity", 5)

    # Collect data rows (row 2 onward) and which are empty
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows_data.append(list(row))
    # Keep only non-empty rows
    kept = [r for r in rows_data if not is_empty_row(r, id_col, name_col)]

    # Assign popularity 1, 2, 3, ...
    for i, row in enumerate(kept):
        while len(row) <= pop_col:
            row.append(None)
        row[pop_col] = i + 1

    # Clear sheet from row 2 to end, then write back
    max_row = ws.max_row
    if max_row >= 2:
        ws.delete_rows(2, max_row - 1)
    for i, row in enumerate(kept):
        for j, val in enumerate(row):
            ws.cell(row=2 + i, column=j + 1, value=val)

    # Recreate table (ref may have changed)
    new_max = len(kept) + 1
    max_col = len(HEADERS)
    ref = "A1:{}".format(get_column_letter(max_col)) + str(new_max)
    # Remove existing table if any
    if ws.tables:
        del ws.tables["ArtistsTable"]
    tab = Table(displayName="ArtistsTable", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    wb.save(EXCEL_PATH)
    print("Excel: removed empty rows, set Popularity to 1..{}".format(len(kept)))
    wb.close()

    # Update JSON from Excel
    from excel_to_json import import_excel_to_json
    import_excel_to_json(EXCEL_PATH, JSON_PATH)
    print("JSON updated from Excel.")


if __name__ == "__main__":
    main()
