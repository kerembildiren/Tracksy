"""
Export artists_raw.json to an Excel file for easy manual editing.
Table format with header row, frozen pane, and sensible column widths.
Genres are stored as semicolon-separated in one cell (e.g. "Rap" or "Rap; Rock").
"""
import json
import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

# Column order and display headers (key in JSON, header for Excel)
COLUMNS = [
    ("id", "ID"),
    ("name", "Name"),
    ("gender", "Gender"),
    ("genres", "Genre(s)"),
    ("popularity", "Popularity"),
    ("debut", "Debut"),
    ("nationality", "Nationality"),
    ("group_size", "Group size"),
    ("spotify_monthly_streams", "Spotify monthly streams"),
    ("image_url", "Image URL"),
    ("top_track_name", "Top track name"),
    ("top_track_uri", "Top track URI"),
    ("top_track_id", "Top track ID"),
]

GENRE_SEP = "; "  # Semicolon to separate multiple genres in one cell (no comma in genre names)


def artist_to_row(artist):
    """Convert one artist dict to a list of cell values in COLUMNS order."""
    row = []
    for key, _ in COLUMNS:
        val = artist.get(key)
        if key == "genres":
            if isinstance(val, list):
                val = GENRE_SEP.join(str(g).strip() for g in val if g)
            else:
                val = "" if val is None else str(val)
        elif val is None:
            val = ""
        else:
            val = val
        row.append(val)
    return row


def export_json_to_excel(json_path, excel_path=None):
    if excel_path is None:
        excel_path = os.path.join(os.path.dirname(json_path), "artists_raw.xlsx")

    with open(json_path, "r", encoding="utf-8") as f:
        artists = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Artists"

    # Headers
    headers = [h for _, h in COLUMNS]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, artist in enumerate(artists, 2):
        for col_idx, val in enumerate(artist_to_row(artist), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Excel table (for filtering/sorting)
    max_row = len(artists) + 1
    max_col = len(COLUMNS)
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    tab = Table(displayName="ArtistsTable", ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Column widths (one per column)
    widths = [26, 28, 10, 18, 10, 8, 14, 12, 18, 50, 32, 50, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(excel_path)
    print(f"Exported {len(artists)} artists to {excel_path}")
    return excel_path


if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, "output", "artists_raw.json")
    export_json_to_excel(json_path)
