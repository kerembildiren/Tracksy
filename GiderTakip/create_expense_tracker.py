# -*- coding: utf-8 -*-
"""
Creates a Turkish expense tracker Excel file.
Sheet 1: Gelir (left) + Gider (right) side by side, same rows. Sheet 2: Grafik (charts only, formulas from sheet 1).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

GELIR_HEADERS = ["Tarih", "Kategori", "Tutar (₺)", "Açıklama"]
GIDER_HEADERS = ["Tarih", "Kategori", "Tutar (₺)", "Açıklama"]

GELIR_CATEGORIES = ["Maaş", "Diğer"]
GIDER_CATEGORIES = [
    "Kira",
    "YK Kredi Kartı",
    "Ziraat Kredi Kartı",
    "Aidat",
    "Euro",
    "İnternet Faturası",
    "Diğer",
]

# Gelir: columns A-D, rows 1=title, 2=header, 3..252=data
# Gider: columns F-I (same row numbers)
# Summary: column K
GELIR_COLS = {"tarih": 1, "kategori": 2, "tutar": 3, "aciklama": 4}
GIDER_COLS = {"tarih": 6, "kategori": 7, "tutar": 8, "aciklama": 9}
DATA_START_ROW = 3
DATA_END_ROW = 252
SUMMARY_COL = 11  # K


def style_header(ws, row, start_col, num_cols=4):
    """Style header row for a block of columns."""
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for c in range(start_col, start_col + num_cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_data_sheet(wb):
    """One sheet: Gelir (left) and Gider (right) side by side, same rows. Summary in column K."""
    ws = wb.active
    ws.title = "Gelir_Gider"

    # ---- GELİR block (columns A-D) ----
    ws.cell(row=1, column=1, value="GELİR")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    for col, h in enumerate(GELIR_HEADERS, 1):
        ws.cell(row=2, column=col, value=h)
    style_header(ws, 2, 1)

    dv_gelir = DataValidation(
        type="list",
        formula1=f'"{",".join(GELIR_CATEGORIES)}"',
        allow_blank=True,
    )
    dv_gelir.error = "Kategori listesinden seçin"
    dv_gelir.prompt = "Gelir kategorisi"
    ws.add_data_validation(dv_gelir)
    dv_gelir.add(f"B{DATA_START_ROW}:B{DATA_END_ROW}")

    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        ws.cell(row=r, column=GELIR_COLS["tarih"]).number_format = "DD.MM.YYYY"
        ws.cell(row=r, column=GELIR_COLS["tutar"]).number_format = "#,##0.00 ₺"

    ws.cell(row=DATA_START_ROW, column=1, value="01.01.2025")
    ws.cell(row=DATA_START_ROW, column=2, value="Maaş")
    ws.cell(row=DATA_START_ROW, column=3, value=0)
    ws.cell(row=DATA_START_ROW, column=4, value="Örnek - silebilirsiniz")

    # ---- GİDER block (columns F-I), same row numbers ----
    ws.cell(row=1, column=6, value="GİDER")
    ws.cell(row=1, column=6).font = Font(bold=True, size=12)
    # Show category list so user sees where graph gets category names from
    ws.cell(row=1, column=7, value="(Kategoriler: Kira, YK Kredi Kartı, Ziraat Kredi Kartı, Aidat, Euro, İnternet Faturası, Diğer)")
    ws.cell(row=1, column=7).font = Font(italic=True, size=9)
    for col, h in enumerate(GIDER_HEADERS, 6):
        ws.cell(row=2, column=col, value=h)
    style_header(ws, 2, 6)

    dv_gider = DataValidation(
        type="list",
        formula1=f'"{",".join(GIDER_CATEGORIES)}"',
        allow_blank=True,
    )
    dv_gider.error = "Kategori listesinden seçin"
    dv_gider.prompt = "Gider kategorisi"
    ws.add_data_validation(dv_gider)
    dv_gider.add(f"G{DATA_START_ROW}:G{DATA_END_ROW}")

    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        ws.cell(row=r, column=GIDER_COLS["tarih"]).number_format = "DD.MM.YYYY"
        ws.cell(row=r, column=GIDER_COLS["tutar"]).number_format = "#,##0.00 ₺"

    ws.cell(row=DATA_START_ROW, column=6, value="01.01.2025")
    ws.cell(row=DATA_START_ROW, column=7, value="Kira")
    ws.cell(row=DATA_START_ROW, column=8, value=0)
    ws.cell(row=DATA_START_ROW, column=9, value="Örnek - silebilirsiniz")

    # ---- Summary (column K): from Gelir and Gider columns on this sheet ----
    ws.cell(row=1, column=SUMMARY_COL, value="Toplam Gelir (₺)")
    ws.cell(row=1, column=SUMMARY_COL).font = Font(bold=True)
    ws.cell(row=2, column=SUMMARY_COL, value=f"=SUM(C{DATA_START_ROW}:C{DATA_END_ROW})")
    ws.cell(row=2, column=SUMMARY_COL).number_format = "#,##0.00 ₺"

    ws.cell(row=3, column=SUMMARY_COL, value="Toplam Gider (₺)")
    ws.cell(row=3, column=SUMMARY_COL).font = Font(bold=True)
    ws.cell(row=4, column=SUMMARY_COL, value=f"=SUM(H{DATA_START_ROW}:H{DATA_END_ROW})")
    ws.cell(row=4, column=SUMMARY_COL).number_format = "#,##0.00 ₺"

    ws.cell(row=5, column=SUMMARY_COL, value="Bakiye / Artan (₺)")
    ws.cell(row=5, column=SUMMARY_COL).font = Font(bold=True)
    ws.cell(row=6, column=SUMMARY_COL, value="=K2-K4")
    ws.cell(row=6, column=SUMMARY_COL).number_format = "#,##0.00 ₺"
    ws.cell(row=6, column=SUMMARY_COL).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 24
    ws.column_dimensions["K"].width = 18

    return ws


def add_grafik_sheet(wb):
    """
    Grafik sheet: only charts. All numbers come from formulas that read Gelir_Gider.
    Category names (Kira, YK Kredi Kartı...) are the same as the Gider dropdown on sheet 1.
    Each row's value = SUMIF: sum of Tutar on sheet 1 where Kategori = that name. So if you
    never enter "YK Kredi Kartı", that row is 0. No data is invented.
    """
    ws = wb.create_sheet("Grafik", 1)
    data_sheet = "Gelir_Gider"
    # Gider on data sheet: Kategori = col G, Tutar = col H, rows DATA_START_ROW..DATA_END_ROW
    gider_kategori_range = f"{data_sheet}!$G${DATA_START_ROW}:$G${DATA_END_ROW}"
    gider_tutar_range = f"{data_sheet}!$H${DATA_START_ROW}:$H${DATA_END_ROW}"

    ws["A1"] = "Grafikler: Tüm değerler yalnızca önceki sayfadaki Gelir/Gider sütunlarına girdiğiniz verilerden hesaplanır."
    ws["A1"].font = Font(italic=True, size=10)
    ws["A2"] = "Gider grafiğindeki kategoriler, Gider sütunundaki açılır listedeki seçeneklerle aynıdır; tutar = o kategoride girdiğiniz toplam."
    ws["A2"].font = Font(italic=True, size=9)

    # Table: same category names as Gider dropdown; value = SUMIF from sheet 1 (only what you entered)
    ws["A4"] = "Kategori"
    ws["B4"] = "Toplam (₺) = Gider sayfasındaki girişlerinizden"
    style_header(ws, 4, 1, 2)
    for i, cat in enumerate(GIDER_CATEGORIES, 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=f'=SUMIF({gider_kategori_range},A{row},{gider_tutar_range})')
        ws.cell(row=row, column=2).number_format = "#,##0.00 ₺"
    last_cat_row = 4 + len(GIDER_CATEGORIES)

    pie = PieChart()
    pie.title = "Gider Dağılımı (Sadece girdiğiniz Gider kayıtları)"
    data = Reference(ws, min_col=2, min_row=5, max_row=last_cat_row)
    labels = Reference(ws, min_col=1, min_row=5, max_row=last_cat_row)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 12
    pie.width = 18
    ws.add_chart(pie, "D4")

    # Bar: Gelir vs Gider totals from sheet 1 (K2 and K4)
    ws["A23"] = "Tür"
    ws["B23"] = "Tutar"
    ws["A24"] = "Gelir"
    ws["A25"] = "Gider"
    ws["B24"] = f"={data_sheet}!K2"
    ws["B25"] = f"={data_sheet}!K4"
    bar = BarChart()
    bar.type = "col"
    bar.title = "Toplam Gelir vs Gider (sayfa 1 toplamları)"
    bar.y_axis.title = "Tutar (₺)"
    data_bar = Reference(ws, min_col=2, min_row=23, max_row=25)
    cats = Reference(ws, min_col=1, min_row=24, max_row=25)
    bar.add_data(data_bar, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 10
    bar.width = 14
    ws.add_chart(bar, "D20")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    return ws


def main():
    out_path = "Gider_Takip.xlsx"
    wb = Workbook()
    add_data_sheet(wb)
    add_grafik_sheet(wb)
    wb.save(out_path)
    print(f"Oluşturuldu: {out_path}")


if __name__ == "__main__":
    main()
